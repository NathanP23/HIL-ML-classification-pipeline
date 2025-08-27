"""
Excel export functionality with RTL support and color formatting
"""
import json
import os
import glob
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from config.settings import ALL_LABELS
from config.paths import PATHS


def _load_manual_labels():
	"""Load manual labels from the latest master file"""
	try:
		master_pattern = os.path.join(PATHS['manual_master'], "TOTAL_MANUAL_LABEL_AT-*_TOTAL_SAMPLE_SIZE_*.json")
		manual_files = glob.glob(master_pattern)
		
		if manual_files:
			latest_file = sorted(manual_files)[-1]
			with open(latest_file, 'r', encoding='utf-8') as f:
				return json.load(f)
	except Exception as e:
		print(f"⚠️  Could not load manual labels: {e}")
	
	return None


def convert_json_to_excel_rtl(json_path, output_path=None, include_manual=False, source_info=None, sort_by_length=False, include_length_column=False):
	"""
	Convert JSON classification results to Excel format with RTL and color formatting
	
	Args:
		json_path: Path to JSON file with classification results
		output_path: Optional output Excel file path
		include_manual: Whether to include manual labels in the export
		source_info: Dict mapping record IDs to their source ('manual', 'model', 'previously_manual')
		sort_by_length: Whether to sort records by text content length (smallest first)
		include_length_column: Whether to include text length as a column in Excel
	
	Returns:
		Path to created Excel file
	"""
	# Load JSON data
	print(f"📂 Loading: {json_path}")
	with open(json_path, 'r', encoding='utf-8') as f:
		data = json.load(f)
	
	# Include manual labels if needed
	if include_manual:
		manual_data = _load_manual_labels()
		if manual_data:
			print(f"📋 Including {len(manual_data)} manual labels")
			data.extend(manual_data)
			
			# Update source info for manual records
			if source_info is None:
				source_info = {}
			for record in manual_data:
				source_info[record['id']] = 'manual'
	
	print(f"✅ Loaded {len(data)} total records")
	
	# Convert to DataFrame with source tracking
	rows = []
	for record in data:
		row = {
			'id': record.get('id', ''),
			'text_content': record.get('text_content', '')
		}
		
		# Add text length column if requested
		if include_length_column:
			row['text_length'] = len(record.get('text_content', ''))
		
		# Add category values (0/1)
		for category in ALL_LABELS:
			row[category] = record.get(category, 0)
		
		# Add classification source column
		record_id = record.get('id', '')
		if source_info and record_id in source_info:
			row['classification_source'] = source_info[record_id]
		else:
			row['classification_source'] = 'model'  # Default for bulk results
		
		rows.append(row)
	
	df = pd.DataFrame(rows)
	
	# Sort by text length if requested
	if sort_by_length:
		print("🔤 Sorting records by text length (shortest first)...")
		if not include_length_column:
			# Add temporary length column for sorting only
			df['temp_length'] = df['text_content'].str.len()
			df = df.sort_values('temp_length', ascending=True).drop('temp_length', axis=1)
		else:
			# Use existing length column for sorting
			df = df.sort_values('text_length', ascending=True)
		print(f"✅ Sorted {len(df)} records by text length")
	
	# Generate output filename if not provided
	if output_path is None:
		base_name = os.path.splitext(os.path.basename(json_path))[0]
		include_suffix = "_with_manual" if include_manual else ""
		sort_suffix = "_sorted_by_length" if sort_by_length else ""
		length_suffix = "_with_length_col" if include_length_column else ""
		output_path = f"{base_name}{include_suffix}{sort_suffix}{length_suffix}_excel_export.xlsx"
	
	# Save to Excel with RTL and enhanced formatting
	print(f"💾 Saving to: {output_path}")
	with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
		df.to_excel(writer, sheet_name='Classifications', index=False)
		
		# Get worksheet for formatting
		ws = writer.sheets['Classifications']
		
		# Set RTL (Right-to-Left) reading order
		ws.sheet_view.rightToLeft = True

		# Determine column layout based on included columns
		text_col_index = 2  # B column (1-indexed)
		length_col_index = 3 if include_length_column else None  # C column if included
		categories_start_col = 4 if include_length_column else 3  # Categories start column
		
		# Add freeze panes: freeze after text_content column
		freeze_col_letter = get_column_letter(categories_start_col)
		ws.freeze_panes = f'{freeze_col_letter}2'
		
		# Column widths
		ws.column_dimensions['A'].width = 12   # ID column
		ws.column_dimensions['B'].width = 100  # text column
		if include_length_column:
			ws.column_dimensions['C'].width = 10   # text length column
		
		# Color fills and alignments
		green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')     # Light green
		red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')      # Light red
		yellow_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')   # Light yellow for source
		rtl_alignment = Alignment(horizontal='right', readingOrder=2)  # RTL alignment
		center_alignment = Alignment(horizontal='center')
		
		# Category columns - format and color
		for i, category in enumerate(ALL_LABELS):
			col_letter = get_column_letter(i + categories_start_col)  # Dynamic start based on layout
			ws.column_dimensions[col_letter].width = 8
			
			# Color code the category values (starting from row 2, after header)
			for row_idx in range(2, len(data) + 2):  # +2 because Excel is 1-indexed and we have header
				cell = ws[f'{col_letter}{row_idx}']
				cell_value = cell.value
				cell.alignment = center_alignment  # Center align category values
				
				if cell_value == 1:
					cell.fill = green_fill
				elif cell_value == 0:
					cell.fill = red_fill
		
		# Classification source column (last column)
		source_col_letter = get_column_letter(len(ALL_LABELS) + categories_start_col)  # After all categories
		ws.column_dimensions[source_col_letter].width = 20
		
		# Color code source column
		for row_idx in range(2, len(data) + 2):
			cell = ws[f'{source_col_letter}{row_idx}']
			cell.fill = yellow_fill
			cell.alignment = center_alignment
			
			# Add font styling based on source
			if cell.value == 'manual':
				cell.font = Font(bold=True, color='008000')  # Bold green for manual
			elif cell.value == 'previously_manual':
				cell.font = Font(bold=True, color='0066CC')  # Bold blue for previously manual
			else:
				cell.font = Font(color='666666')  # Gray for model
		
		# Apply RTL alignment to text content column (column B)
		for row_idx in range(2, len(data) + 2):  # Start from row 2, after header
			cell = ws[f'B{row_idx}']
			cell.alignment = rtl_alignment
		
		# Format text length column if included (column C)
		if include_length_column:
			for row_idx in range(2, len(data) + 2):
				cell = ws[f'C{row_idx}']
				cell.alignment = center_alignment
		
		# Make header row bold and styled
		header_font = Font(bold=True, size=11)
		header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')  # Light gray
		
		for col in range(1, len(df.columns) + 1):
			col_letter = get_column_letter(col)
			header_cell = ws[f'{col_letter}1']
			header_cell.font = header_font
			header_cell.fill = header_fill
			header_cell.alignment = center_alignment
	
	print(f"🎉 Excel created successfully!")
	total_cols = len(ALL_LABELS) + 3  # id, text_content, classification_source
	if include_length_column:
		total_cols += 1  # Add text_length column
	print(f"📊 Format: {len(data)} rows × {total_cols} columns")
	if sort_by_length:
		print(f"🔤 Sorted: Records arranged by text length (shortest first)")
	if include_length_column:
		print(f"📏 Length: Text length column included")
	print(f"📁 File: {output_path}")
	
	return output_path


def show_excel_export_summary(data, category_cols):
	"""Show summary statistics of the classifications"""
	print(f"\n📊 CLASSIFICATION SUMMARY:")
	print(f"{'Category':<35} {'Count':<8} {'Percentage':<12}")
	print("-" * 60)
	
	total_records = len(data)
	
	for category in category_cols:
		count = sum(1 for record in data if record.get(category, 0) == 1)
		percentage = (count / total_records) * 100
		category_display = category.replace('_', ' ').title()[:34]
		print(f"{category_display:<35} {count:<8} {percentage:<12.1f}%")
	
	# Records with no classifications
	no_classifications = sum(1 for record in data 
						   if all(record.get(cat, 0) == 0 for cat in category_cols))
	print(f"\nRecords with no classifications: {no_classifications}")
	
	# Average classifications per record
	avg_classifications = sum(sum(record.get(cat, 0) for cat in category_cols) 
							for record in data) / len(data)
	print(f"Average classifications per record: {avg_classifications:.2f}")


def export_latest_bulk_results(include_manual_labels=False, sort_by_length=False, include_length_column=False):
	"""Export latest bulk classification results with optional manual labels"""
	title = "EXPORT LATEST BULK RESULTS" + (" + MANUAL LABELS" if include_manual_labels else "")
	print(f"🚀 {title} TO EXCEL")
	print("=" * 60)
	
	# Find latest bulk classification file
	bulk_pattern = os.path.join(PATHS['base_output_dir'], "bulk_classification", "final", "*.json")
	bulk_files = glob.glob(bulk_pattern)
	
	if not bulk_files:
		print(f"❌ No bulk classification files found")
		return None
	
	latest_file = sorted(bulk_files)[-1]
	print(f"📁 Using latest bulk file: {os.path.basename(latest_file)}")
	
	try:
		# Create source info for bulk + manual combination
		source_info = {}
		
		if include_manual_labels:
			# Load manual labels to mark them appropriately
			manual_data = _load_manual_labels()
			if manual_data:
				print(f"📋 Will include {len(manual_data)} manual labels")
				for record in manual_data:
					source_info[record['id']] = 'manual'
		
		# Check if this is a bulk file with manual labels (filename contains WITH_MANUAL)
		if "WITH_MANUAL" in os.path.basename(latest_file):
			print(f"🔄 Detected bulk file with re-classified manual labels")
			# Load manual IDs to mark re-classified records
			manual_data = _load_manual_labels()
			if manual_data:
				manual_id_set = {record['id'] for record in manual_data}
				# Mark records that were re-classified
				with open(latest_file, 'r', encoding='utf-8') as f:
					bulk_data_check = json.load(f)
				for record in bulk_data_check:
					if record['id'] in manual_id_set:
						source_info[record['id']] = 'previously_manual'
		
		# Convert to Excel with enhanced features
		excel_path = convert_json_to_excel_rtl(
			latest_file, 
			include_manual=include_manual_labels,
			source_info=source_info,
			sort_by_length=sort_by_length,
			include_length_column=include_length_column
		)
		
		# Load data for summary
		with open(latest_file, 'r', encoding='utf-8') as f:
			bulk_data = json.load(f)
		
		# Show summary
		total_records = len(bulk_data)
		if include_manual_labels:
			manual_data = _load_manual_labels()
			if manual_data:
				total_records += len(manual_data)
		
		print(f"\n🎉 BULK RESULTS EXPORTED TO EXCEL!")
		print(f"   📁 Excel file: {excel_path}")
		print(f"   📊 Total records: {total_records}")
		if sort_by_length:
			print(f"   🔤 Sorting: Records sorted by text length (shortest first)")
		if include_length_column:
			print(f"   📏 Length: Text length column included")
		print(f"   🎨 Features: RTL layout, freeze panes, source tracking")
		print(f"   🎨 Color coding: Green=1, Red=0, Yellow=source column")
		
		if include_manual_labels:
			print(f"   📋 Includes manual labels for comparison")
		
		return excel_path
		
	except Exception as e:
		print(f"❌ Error: {e}")
		return None


def export_manual_labels(sort_by_length=False, include_length_column=False):
	"""Export manual labels to Excel"""
	print("🚀 EXPORT MANUAL LABELS TO EXCEL")
	print("=" * 40)
	
	# Find latest manual label file
	from config.paths import PATHS
	import glob
	
	manual_pattern = os.path.join(PATHS['manual_master'], "TOTAL_MANUAL_LABEL_AT-*_TOTAL_SAMPLE_SIZE_*.json")
	manual_files = glob.glob(manual_pattern)
	
	if not manual_files:
		print(f"❌ No manual label files found")
		return None
	
	latest_file = sorted(manual_files)[-1]
	print(f"📁 Using latest file: {os.path.basename(latest_file)}")
	
	try:
		# Convert to Excel
		excel_path = convert_json_to_excel_rtl(latest_file, sort_by_length=sort_by_length, include_length_column=include_length_column)
		
		# Load data for summary
		with open(latest_file, 'r', encoding='utf-8') as f:
			data = json.load(f)
		
		show_excel_export_summary(data, ALL_LABELS)
		
		print(f"\n🎉 MANUAL LABELS EXPORTED TO EXCEL!")
		print(f"   📁 Excel file: {excel_path}")
		if sort_by_length:
			print(f"   🔤 Sorting: Records sorted by text length (shortest first)")
		if include_length_column:
			print(f"   📏 Length: Text length column included")
		print(f"   🎨 Features: RTL layout, color coding (green=1, red=0)")
		
		return excel_path
		
	except Exception as e:
		print(f"❌ Error: {e}")
		return None